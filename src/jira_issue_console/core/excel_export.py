"""Functions for exporting data to Excel format."""

from typing import Any, Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore[assignment,misc]


def export_to_excel(
    rows: List[Dict[str, Any]], filepath: str, fieldnames: List[str]
) -> None:
    """Export data rows to Excel format.

    Args:
        rows: List of dicts where each dict has the column data
        filepath: Path to the Excel file to create
        fieldnames: List of field names in order
    """
    if Workbook is None:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    # Write headers
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        ws.cell(row=1, column=col_idx, value=fieldname)

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            value = row.get(fieldname, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns (approximate)
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(fieldname)
        for row in rows:
            cell_value = str(row.get(fieldname, ""))
            max_length = max(max_length, len(cell_value))
        # Set column width (with some padding)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(filepath)


def get_file_extension(format_type: str) -> str:
    """Get file extension for the given format type.

    Args:
        format_type: Either 'csv' or 'excel'

    Returns:
        File extension including the dot (e.g., '.csv' or '.xlsx')
    """
    if format_type == "excel":
        return ".xlsx"
    return ".csv"
